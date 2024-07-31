import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook

url = "https://gosstroy.gov.kg/ru/kyzmat/1/reestr-legal-buildings"
response = requests.get(url)
html_content = response.text

soup = BeautifulSoup(html_content, "html.parser")

table = soup.find("table")

if table is None:
    print("Таблица не найдена.")
else:
    rows = table.find_all("tr")

    wb = Workbook()
    ws = wb.active
    ws.title = "Таблица данных"

    headers = rows[0].find_all("th")
    for j, header in enumerate(headers, start=1):
        ws.cell(row=1, column=j, value=header.text.strip())

    for i, row in enumerate(rows[1:], start=2):
        cells = row.find_all("td")
        for j, cell in enumerate(cells, start=1):
            ws.cell(row=i, column=j, value=cell.text.strip())

    wb.save("gosstroy.xlsx")
