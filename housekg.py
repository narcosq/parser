import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook

def extract_data_from_page(soup, ws, start_row, seen_titles):
    main_block = soup.find("div", class_="buildings-table")
    if not main_block:
        print("Не удалось найти блок с объявлениями.")
        return start_row

    listings = main_block.find_all("div", class_="building-item")
    print(f"Найдено {len(listings)} объявлений на странице.")

    for listing in listings:
        try:
            title = listing.find("p", class_="title").text.strip()
        except AttributeError:
            title = "N/A"

        if title in seen_titles:
            print(f"Объявление '{title}' уже записано, пропуск.")
            continue
        seen_titles.add(title)

        try:
            status = listing.find("div", class_="status-label").text.strip()
        except AttributeError:
            status = "N/A"

        try:
            address = listing.find("div", class_="address-building").text.strip()
        except AttributeError:
            address = "N/A"

        try:
            storeys = listing.find("div", class_="description-listing").text.strip()
        except AttributeError:
            storeys = "N/A"

        try:
            builder = listing.find("div", class_="builder").text.strip()
        except AttributeError:
            builder = "N/A"

        ws.cell(row=start_row, column=1, value=title)
        ws.cell(row=start_row, column=2, value=address)
        ws.cell(row=start_row, column=3, value=storeys)
        ws.cell(row=start_row, column=4, value=builder)
        ws.cell(row=start_row, column=5, value=status)

        start_row += 1

    return start_row

wb = Workbook()
ws = wb.active
ws.title = "Объявления"
headers = ["Наименование объекта", "Адрес объекта", "Количество этажей", "Заказчик", "Статус"]
for i, header in enumerate(headers, start=1):
    ws.cell(row=1, column=i, value=header)

seen_titles = set()

page = 1
start_row = 2
max_pages = 72

while page <= max_pages:
    url = f"https://www.house.kg/jilie-kompleksy?page={page}"
    response = requests.get(url)
    soup = BeautifulSoup(response.text, "html.parser")

    start_row = extract_data_from_page(soup, ws, start_row, seen_titles)

    page += 1

excel_file = "housekg.xlsx"
wb.save(excel_file)
print(f"Файл сохранен как {excel_file}")
